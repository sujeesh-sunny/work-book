from openpyxl import Workbook
from openpyxl.styles import Font


wb = Workbook()

ws1 = wb.active
ws1.title = 'SHEET 1'
ws1['a1'] = 'sujeesh'
ws1['a2'] = 'rotech'
ws1['b1'] = 'rohit'

bold_txt = Font(bold=True)


ws2=wb.create_sheet('SHEET 2')
ws2['a1'] = 'sujeesh'
ws2['a2'] = 'rotech'

ws3=wb.create_sheet()
ws3['a1'] = 'sujeesh'

ws1['a1'].font=bold_txt

wb.save('wb1.xlsx')


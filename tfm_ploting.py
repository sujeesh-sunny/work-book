from flask import Flask
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)

def item_parser(category, search_term, site, wb):
    with open('configs/competitor.json', 'r') as file:
        competitor = json.load(file)[site]     
    URL = f"{competitor['store_api']}/{category}/{search_term}"
    HEADERS = { 
        'Accept-Language': "en-US,en;q=0.9,hi;q=0.8",
        'User-Agent': "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36",
        'Cookie': competitor['cookie']
    }
    response = requests.get(URL, headers=HEADERS)
    data = response.json()
    if 'items' in data:
        items = data['items']
        for item in items:
            categories = item['categories']
            for category_dict in categories:            
                if category in category_dict['name'].lower() and search_term in item['name'].lower():
                    item_found={
                                    "name":item['name'],
                                    "price":item['base_price'],
                                    }
                    status=200
                    break   
        return ({"status": status, "item": item_found})
    else:
        return ({"data": data})

@app.route('/api/v1/getitem/<site>/<category>/<search_term>')   
def getitem(site, category, search_term):
    wb = Workbook()
    ws = wb.active
    ws.title = 'SHEET 1'

    bold_txt = Font(bold=True)
    ws['A1'].font = bold_txt
    ws['B1'].font = bold_txt

    result = item_parser(category, search_term, site, wb)

    if result['status'] == 200:
        item = result['item']
        ws['A1'] = item['name']
        ws['B1'] = item['price']
        wb.save('tfm_plotting.xlsx')
        return {"status": 200, "message": "Data plotted and saved successfully"}
    else:
        return result

if __name__ == '__main__':
    app.run(debug=True, port=5000)

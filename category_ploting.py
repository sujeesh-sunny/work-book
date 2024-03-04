from flask import Flask
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)

def item_parser(category, search_term, site, wb, row):
    with open('configs/competitor.json', 'r') as file:
        competitor = json.load(file).get(site)
        if not competitor:
            return {"status": 404, "message": f"Site '{site}' not found in configuration"}
        
    URL = f"{competitor['store_api']}/{category}/{search_term}"
    HEADERS = { 
        'Accept-Language': "en-US,en;q=0.9,hi;q=0.8",
        'User-Agent': "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36",
        'Cookie': competitor['cookie']
    }
    response = requests.get(URL, headers=HEADERS)
    data = response.json()
    
    item_found = None  # Initialize item_found variable
    
    if 'items' in data:
        items = data['items']
        for item in items:
            categories = item.get('categories', [])
            for category_dict in categories:
                if category in category_dict.get('name', '').lower() and search_term in item.get('name', '').lower():
                    item_found = {
                        "name": item.get('name'),
                        "price": item.get('base_price'),
                        "category": category_dict.get('name'),
                        "site": site
                    }
                    break  # Exit loop once item is found
    
    if item_found is None:  # Check if item_found is still None (i.e., not found)
        return {"status": 404, "message": "Item not found"}

    ws = wb.active
    ws['A' + str(row)] = item_found['site']
    ws['B' + str(row)] = item_found['name']
    ws['C' + str(row)] = item_found['price']
    ws['D' + str(row)] = item_found['category']
    return {"status": 200, "message": "Data written successfully"}


@app.route('/api/v1/getitem/<site>/<category>/<search_term>')   
def getitem(site, category, search_term):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparison'

    bold_txt = Font(bold=True)
    ws['A1'].font = bold_txt
    ws['B1'].font = bold_txt
    ws['C1'].font = bold_txt
    ws['D1'].font = bold_txt

    ws['A1'] = 'Site'
    ws['B1'] = 'Name'
    ws['C1'] = 'Price'
    ws['D1'] = 'Category'

    row = 2
    sites = ['sprouts', 'wegmans', 'tfm']
    for site in sites:
        result = item_parser(category, search_term, site, wb, row)
        if result['status'] != 200:
            return result
        row += 1
    
    wb.save('comparison.xlsx')
    return {"status": 200, "message": "Data compared and saved successfully"}

if __name__ == '__main__':
    app.run(debug=True, port=5000)
